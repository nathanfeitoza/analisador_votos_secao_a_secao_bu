from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def create_table(wb, title):
  ws = wb.active
  ws.title = title # Table Title
  
  return ws

def get_thin_border():
  thin_border = Border(
      left=Side(style='thin'),
      right=Side(style='thin'),
      top=Side(style='thin'),
      bottom=Side(style='thin')
  )
  
  return thin_border

def get_fills():
  return {
    'header_fill': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    'lighter_fill': PatternFill(start_color="DFDFDF", end_color="DFDFDF", fill_type="solid")
  }


def create_header(ws, title):
  title_font = Font(bold=True, size=14)
  header_font = Font(bold=True, color="000000")
  alignment_center = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('A1:D1') 
  ws['A1'] = title
  ws['A1'].font = title_font
  ws['A1'].alignment = alignment_center
  
  headers = ['Local de Votação', 'Seção', 'Quantidade de Votos', 'Total de Votos no Local']
  ws.append(headers)
  
  fills = get_fills()
  border = get_thin_border()
  
  for cell in ws[2]:
      cell.font = header_font
      cell.alignment = alignment_center
      cell.border = border
      cell.fill = fills['header_fill']
      
  return ws

def add_data(ws, data):
  fills = get_fills()
  border = get_thin_border()
  alignment_center = Alignment(horizontal='center', vertical='center')

  def add_empty_row(ws):
      ws.append(["", "", "", ""])
      for cell in ws[len(ws['A'])]:
          cell.fill = fills['lighter_fill']
          cell.border = border

  for i, row in enumerate(data):
      if len(row[0]) != 0 and i != 0:
          add_empty_row(ws)
      ws.append(row)
      for cell in ws[len(ws['A'])]:
          cell.border = border
          cell.alignment = alignment_center

  return ws


def add_total(ws, data):
  fills = get_fills()
  border = get_thin_border()
  alignment_center = Alignment(horizontal='center', vertical='center')
  
  data_add = [
    ["", "", "", ""],
    data
  ]

  for i, row in enumerate(data_add):
    ws.append(row)
    
    for cell in ws[len(ws['A'])]:
        cell.border = border
        cell.alignment = alignment_center
        if i != 0:
          cell.fill = fills['lighter_fill']
      
  
  return ws

def style_columns(ws):
  column_widths = [30, 10, 20, 25]
  for i, width in enumerate(column_widths, start=1):
      ws.column_dimensions[chr(64 + i)].width = width


def save(wb, file_name):
  output_path = f"./output/{file_name}.xlsx"
  wb.save(output_path)
  print(f"Tabela criada e salva como {output_path}")


def build_excel_table(table_tab_title, table_title, file_name, data, data_total):
  wb = Workbook()
  table = create_table(wb, table_tab_title)
  header = create_header(table, table_title)
  table_data = add_data(header, data)
  table_total = add_total(table_data, data_total)
  
  style_columns(table_total)
  
  return save(wb, file_name)
  